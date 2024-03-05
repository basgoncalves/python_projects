import docx
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter
from tkinter import filedialog

def copy_paragraphs_from_word():
    # Ask user to select a Word document
    tkinter.Tk().withdraw()
    word_file_path = filedialog.askopenfilename()

    # Load the Word document
    if word_file_path.endswith(".docx"):
        doc = docx.Document(word_file_path)
    elif word_file_path.endswith(".rtf.doc"):
        # Use appropriate methods to load the RTF document
       with open(word_file_path) as infile:
            for line in infile:
                print(line)
        
    else:
        print("Unsupported file format. Please provide a .docx or .rtf.doc file.")
        return
    
    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Set font style for unbold and capitals
    font = Font(bold=False)

    # Copy paragraphs to Excel
    for paragraph in doc.paragraphs:
        # Get the text from the paragraph
        text = paragraph.text
        if paragraph.runs:
            for run in paragraph.runs:
                if run.bold:
                    text = text.replace(run.text, run.text.upper())

        # Add the text to the Excel sheet
        sheet.append([text])

        # Apply font style to the cell
        sheet.cell(row=sheet.max_row, column=1).font = font

    # Save the Excel file in the same folder as the Word document
    excel_file_path = word_file_path.replace(".docx", ".xlsx")
    workbook.save(excel_file_path)

    print("Paragraphs copied to Excel successfully!")

# Call the function
copy_paragraphs_from_word()
